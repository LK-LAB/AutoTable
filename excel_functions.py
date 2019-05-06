import openpyxl as pyxl
import os
from operator import eq
from dictionary import features

pxl_row = 1
pxl_col = 'A'
locate = 'A1'


def searching(element_data, filename):
    temp = filename[2]
    return temp


def location(pxl_row, pxl_col):
    locate = pxl_col + str(pxl_row)
    return locate


def move_row(pxl_row):
    pxl_row + 1


def move_col(pxl_col):
    if pxl_col[len(pxl_col)-1] != 'Z':
        pxl_col[len(pxl_col)-1] = ord(int(pxl_col[len(pxl_col)-1])+1)
    else:
        pxl_col + 'A'


def loading():
    import glob
    return glob.glob("./ElementData/*")

def cell_add(cell_value, feature, data):
    return str(cell_value).replace('None', '') + feature + ': ' + data


#Name : 2, AtomicNumber : 5, Valence : 69, SuperconductingPoint : 64,SpaceGroupName : 60, SpaceGroupNumber : 61
#Period : 49, Group : 30


def auto_table(savefilename, elementlist, featurelist):
    tempwb = pyxl.Workbook()
    tempworksheet = tempwb.active
    tempworksheet.title = 'Table'
    error='Missing["NotApplicable"]\n'
    for element in elementlist:
        tempelement = open(element, 'rt', encoding='UTF8')
        tempdata = tempelement.readlines()
        print("Start %s" %tempdata[0])
        if eq(tempdata[29], error):
            if int(tempdata[48]) == 6: #period is 6th
                tempworksheet.cell(row=(int(tempdata[48]) + 3), column=(int(tempdata[4]) - (57-4))).value = str(tempworksheet.cell(row=(int(tempdata[48]) + 3), column=(int(tempdata[4]) - (57-4))).value) + tempdata[0] + '\r\n'
            else:
                tempworksheet.cell(row=(int(tempdata[48]) + 3), column=(int(tempdata[4]) - (89-4))).value = str(tempworksheet.cell(row=(int(tempdata[48]) + 3), column=(int(tempdata[4]) - (89-4))).value) + tempdata[0] + '\r\n'
        else:
            tempworksheet.cell(row=(int(tempdata[48]) + 1), column=(int(tempdata[29]) + 1)).value = str(tempworksheet.cell(row=(int(tempdata[48]) + 1), column=(int(tempdata[29]) + 1)).value) + str(tempdata[0]) + '\r\n'
    
        for feature in featurelist:
            if tempdata[29] != 'Missing["NotApplicable"]\n':
                tempworksheet.cell(row=(int(tempdata[48]) + 1), column=(int(tempdata[29]) + 1)).value = cell_add( tempworksheet.cell(row=(int(tempdata[48]) + 1), column=(int(tempdata[29]) + 1)).value,feature, tempdata[features[feature]-1])
            else:
                if int(tempdata[48]) == 6: #period is 6th
                    tempworksheet.cell(row=(int(tempdata[48]) + 3), column=(int(tempdata[4]) - (57-4))).value = cell_add(tempworksheet.cell(row=(int(tempdata[48]) + 3), column=(int(tempdata[4]) - (57-4))).value, feature, tempdata[features[feature]-1])
                else:
                    tempworksheet.cell(row=(int(tempdata[48]) + 3), column=(int(tempdata[4]) - (89-4))).value = cell_add( tempworksheet.cell(row=(int(tempdata[48]) + 3), column=(int(tempdata[4]) - (89-4))).value, feature, tempdata[features[feature]-1])
        tempelement.close()
    tempwb.save(filename=("%s.xlsx"%savefilename))
    print("Finish")
